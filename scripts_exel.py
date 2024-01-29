from xlrd import open_workbook

reader = open_workbook("resources/file_example_XLS_10.xls")

sheet = reader.sheet_by_index(0)
print(sheet.nrows)
print(sheet.ncols)
print(sheet.cell_value(4, 4))
for nx in range(sheet.nrows):
    print(sheet.row(nx))

from openpyxl import load_workbook

workbook = load_workbook("resources/file_example_XLSX_50.xlsx")

sheet = workbook.active

print(sheet.cell(2, 3).value)
